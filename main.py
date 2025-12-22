#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Conversor CNAB para Excel - UY3
"""

import os
import sys
import re
from datetime import datetime

# Detectar pasta onde o .exe está (não a pasta temporária)
if getattr(sys, 'frozen', False):
    # Executando como .exe
    PASTA_BASE = os.path.dirname(sys.executable)
else:
    # Executando como script Python
    PASTA_BASE = os.path.dirname(os.path.abspath(__file__))

PASTA_CNAB = os.path.join(PASTA_BASE, "CNABs")
PASTA_SAIDA = os.path.join(PASTA_BASE, "Convertidos")


def banner():
    print("""
╔══════════════════════════════════════════════════════════════════╗
║                                                                  ║
║   ██████╗███╗   ██╗ █████╗ ██████╗                              ║
║  ██╔════╝████╗  ██║██╔══██╗██╔══██╗                             ║
║  ██║     ██╔██╗ ██║███████║██████╔╝                             ║
║  ██║     ██║╚██╗██║██╔══██║██╔══██╗                             ║
║  ╚██████╗██║ ╚████║██║  ██║██████╔╝                             ║
║   ╚═════╝╚═╝  ╚═══╝╚═╝  ╚═╝╚═════╝                              ║
║                                                                  ║
║           CONVERSOR DE ARQUIVOS CNAB - UY3                       ║
║                                                                  ║
╚══════════════════════════════════════════════════════════════════╝
""")


def garantir_pastas():
    """Cria as pastas necessárias se não existirem"""
    if not os.path.exists(PASTA_CNAB):
        os.makedirs(PASTA_CNAB)
        print(f"📁 Pasta 'CNABs' criada em: {PASTA_CNAB}")
    
    if not os.path.exists(PASTA_SAIDA):
        os.makedirs(PASTA_SAIDA)
        print(f"📁 Pasta 'Convertidos' criada em: {PASTA_SAIDA}")


def listar_arquivos_cnab():
    """Lista arquivos CNAB na pasta"""
    arquivos = []
    for f in os.listdir(PASTA_CNAB):
        caminho = os.path.join(PASTA_CNAB, f)
        if os.path.isfile(caminho):
            arquivos.append(caminho)
    return arquivos


def parse_valor(valor_str):
    """Converte valor CNAB (centavos) para número float"""
    try:
        valor_str = valor_str.strip()
        if not valor_str:
            return 0.0
        # Limpar caracteres não numéricos e converter
        valor_limpo = re.sub(r'\D', '', valor_str)
        if not valor_limpo or valor_limpo == '0':
            return 0.0
        valor = int(valor_limpo) / 100.0
        return valor
    except:
        return 0.0


def numero_para_valor(numero):
    """Converte número para formato 'R$ X.XXX,XX'"""
    return f"R$ {numero:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def formatar_cpf_cnpj(doc):
    """Formata CPF (sempre 11 dígitos)"""
    doc = re.sub(r'\D', '', doc)
    # Pegar apenas os últimos 11 dígitos se tiver mais
    if len(doc) > 11:
        doc = doc[-11:]
    # Garantir que tem 11 dígitos, preenchendo com zeros à esquerda se necessário
    doc = doc.zfill(11)
    if len(doc) == 11:
        return f"{doc[:3]}.{doc[3:6]}.{doc[6:9]}-{doc[9:]}"
    return doc


def formatar_cep(cep):
    """Formata CEP"""
    cep = re.sub(r'\D', '', cep)
    if len(cep) == 8:
        return f"{cep[:5]}-{cep[5:]}"
    return cep


def formatar_telefone(tel):
    """Formata telefone"""
    tel = re.sub(r'\D', '', tel)
    if len(tel) == 11:
        return f"({tel[:2]}) {tel[2:7]}-{tel[7:]}"
    elif len(tel) == 10:
        return f"({tel[:2]}) {tel[2:6]}-{tel[6:]}"
    return tel


def parse_cnab_linha(linha):
    """Parse de uma linha CNAB 400 - Layout UY3/Bradesco"""
    if len(linha) < 400:
        return None
    
    tipo_registro = linha[0:1]
    
    # Só processar registros de detalhe (tipo 1)
    if tipo_registro != '1':
        return None
    
    # ID do título com parcela (posições 108-122) - formato: 2104757146-027
    id_titulo_parcela = linha[108:122].strip()
    
    # Extrair apenas o número do meio (remover prefixo 210 e sufixo -027)
    id_titulo = id_titulo_parcela
    
    # Remover prefixo "210" se existir
    if id_titulo.startswith('210'):
        id_titulo = id_titulo[3:]
    
    # Remover sufixo após hífen (ex: -027)
    if '-' in id_titulo:
        partes = id_titulo.split('-')
        id_titulo = partes[0]
    
    # Nome do cliente (posições 234-274)
    nome = linha[234:274].strip()
    
    # CPF/CNPJ (posições 220-234)
    cpf_cnpj_raw = linha[220:234].strip()
    cpf_cnpj = formatar_cpf_cnpj(cpf_cnpj_raw)
    
    # Valor - buscar padrão: zeros + valor_em_centavos + 457
    valor = 0.0
    # Buscar padrão na linha inteira: múltiplos zeros seguidos de dígitos e terminando em 457
    valor_match = re.search(r'0{5,}(\d+?)457', linha)
    if valor_match:
        valor_centavos = valor_match.group(1)
        valor = int(valor_centavos) / 100.0
    
    # Endereço (posições 274-326)
    endereco = linha[274:326].strip()
    
    # CEP - buscar 8 dígitos na região 310-340
    cep = ""
    cep_region = linha[310:340]
    cep_match = re.search(r'\d{8}', cep_region)
    if cep_match:
        cep = formatar_cep(cep_match.group())
    
    # Email - buscar padrão de email na região 326-385 (evitar CEP no início)
    email = ""
    email_region = linha[326:385]
    # Garantir que o email comece com letra, não com número
    email_match = re.search(r'[a-zA-Z][a-zA-Z0-9._%+-]*@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', email_region)
    if email_match:
        email = email_match.group().lower()
    
    # Telefone (posições 382-394)
    telefone_raw = linha[382:394].strip()
    telefone = formatar_telefone(telefone_raw)
    
    return {
        'nome': nome,
        'cpf_cnpj': cpf_cnpj,
        'valor': valor,
        'id_titulo': id_titulo,  # Usar ID limpo (sem prefixo 210 e sem sufixo -027)
        'endereco': endereco,
        'cep': cep,
        'email': email,
        'telefone': telefone
    }


def processar_arquivo(caminho_cnab):
    """Processa um arquivo CNAB e gera Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        print("❌ Erro: biblioteca openpyxl não encontrada")
        return False
    
    # Ler arquivo CNAB
    registros = []
    encodings = ['latin-1', 'utf-8', 'cp1252']
    
    for enc in encodings:
        try:
            with open(caminho_cnab, 'r', encoding=enc) as f:
                linhas = f.readlines()
            break
        except:
            continue
    else:
        print(f"❌ Erro ao ler arquivo: {caminho_cnab}")
        return False
    
    # Processar linhas
    for linha in linhas:
        if len(linha) >= 400:
            reg = parse_cnab_linha(linha)
            if reg:
                registros.append(reg)
    
    if not registros:
        print(f"⚠️  Nenhum registro encontrado em: {os.path.basename(caminho_cnab)}")
        return False
    
    # Agrupar registros por operação (usando id_titulo como chave única)
    operacoes_agrupadas = {}
    for reg in registros:
        chave = reg['id_titulo']  # Usar número da operação como identificador único
        
        if chave not in operacoes_agrupadas:
            # Primeira ocorrência: criar entrada
            operacoes_agrupadas[chave] = {
                'nome': reg['nome'],
                'cpf_cnpj': reg['cpf_cnpj'],
                'quantidade_parcelas': 1,
                'valor_total': reg['valor'],  # Já é número float
                'id_titulo': reg['id_titulo'],
                'endereco': reg['endereco'],
                'cep': reg['cep'],
                'email': reg['email'],
                'telefone': reg['telefone']
            }
        else:
            # Operação já existe: incrementar parcelas e somar valor
            operacoes_agrupadas[chave]['quantidade_parcelas'] += 1
            operacoes_agrupadas[chave]['valor_total'] += reg['valor']  # Já é número float
    
    # Converter dicionário de volta para lista
    registros_agrupados = []
    for operacao in operacoes_agrupadas.values():
        registros_agrupados.append({
            'nome': operacao['nome'],
            'cpf_cnpj': operacao['cpf_cnpj'],
            'parcelas': str(operacao['quantidade_parcelas']),
            'valor': numero_para_valor(operacao['valor_total']),
            'id_titulo': operacao['id_titulo'],
            'endereco': operacao['endereco'],
            'cep': operacao['cep'],
            'email': operacao['email'],
            'telefone': operacao['telefone']
        })
    
    # Criar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados CNAB"
    
    # Cabeçalhos
    headers = ['Nome Cliente', 'CPF/CNPJ', 'Parcelas', 'Valor Total (Soma)', 'Nº da Operação', 'Endereco', 'CEP', 'Email', 'Telefone']
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escrever cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Escrever dados
    for row_idx, reg in enumerate(registros_agrupados, 2):
        valores = [
            reg['nome'],
            reg['cpf_cnpj'],
            reg['parcelas'],
            reg['valor'],
            reg['id_titulo'],
            reg['endereco'],
            reg['cep'],
            reg['email'],
            reg['telefone']
        ]
        for col_idx, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # Ajustar largura das colunas
    larguras = [35, 20, 10, 15, 20, 40, 12, 35, 18]
    for col_idx, largura in enumerate(larguras, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = largura
    
    # Salvar arquivo
    nome_arquivo = os.path.splitext(os.path.basename(caminho_cnab))[0]
    caminho_saida = os.path.join(PASTA_SAIDA, f"{nome_arquivo}.xlsx")
    wb.save(caminho_saida)
    
    print(f"    ✅ Excel gerado: {nome_arquivo}.xlsx ({len(registros_agrupados)} clientes, {len(registros)} parcelas totais)")
    return True


def main():
    banner()
    garantir_pastas()
    
    arquivos = listar_arquivos_cnab()
    
    if not arquivos:
        print(f"\n⚠️  Nenhum arquivo CNAB encontrado.")
        print(f"   Coloque os arquivos CNAB na pasta 'CNABs' e execute novamente.")
        print(f"\n   📁 Pasta CNABs: {PASTA_CNAB}")
        input("\n   Pressione ENTER para sair...")
        return
    
    print(f"\n📂 Encontrado(s) {len(arquivos)} arquivo(s) CNAB para processar\n")
    
    sucessos = 0
    falhas = 0
    
    for i, arquivo in enumerate(arquivos, 1):
        print(f"[{i}/{len(arquivos)}] Processando: {os.path.basename(arquivo)}")
        if processar_arquivo(arquivo):
            sucessos += 1
        else:
            falhas += 1
    
    print(f"\n{'='*50}")
    print(f"RESUMO DO PROCESSAMENTO")
    print(f"  Total de arquivos: {len(arquivos)}")
    print(f"  ✅ Sucessos: {sucessos}")
    if falhas > 0:
        print(f"  ❌ Falhas: {falhas}")
    print(f"\n  📁 Arquivos Excel salvos em: {PASTA_SAIDA}")
    
    input("\n   Pressione ENTER para sair...")


if __name__ == "__main__":
    main()
